"""
Report generators for validation results.

Supports:
  - Rich console output (terminal)
  - HTML report (interactive with search, filter, sort)
  - JSON report
"""

from __future__ import annotations

import json
from collections import Counter
from datetime import datetime
from pathlib import Path

from jinja2 import Template
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from .validators import ValidationResult, ValidationIssue, Severity


# ---------------------------------------------------------------------------
# Console report (using Rich)
# ---------------------------------------------------------------------------

def print_console_report(
    result: ValidationResult,
    severity_filter: str = None,
    rule_filter: str = None,
    limit: int = None,
) -> None:
    """Pretty-print the validation report to the terminal using Rich."""
    from rich.console import Console
    from rich.table import Table
    from rich.panel import Panel
    from rich.columns import Columns

    console = Console()

    # -- Header ---------------------------------------------------------------
    status_color = "green" if result.is_clean else ("red" if result.error_count > 0 else "yellow")
    status_text = "PASS" if result.is_clean else "FAIL"

    console.print()
    console.print(Panel(
        f"[bold]{result.file_name}[/bold]\n"
        f"Rows: {result.total_rows:,}  |  Columns: {result.total_columns}  |  "
        f"Validators: {result.validators_run}\n"
        f"Status: [{status_color} bold]{status_text}[/{status_color} bold]  |  "
        f"Pass Rate: [{status_color}]{result.pass_rate}%[/{status_color}]",
        title="[bold blue]Data Quality Validation Report[/bold blue]",
        border_style="blue",
    ))

    # -- Summary bar ----------------------------------------------------------
    console.print(
        f"  [red]Errors: {result.error_count}[/red]  |  "
        f"[yellow]Warnings: {result.warning_count}[/yellow]  |  "
        f"[cyan]Info: {result.info_count}[/cyan]"
    )
    console.print()

    if not result.issues:
        console.print("[green bold]  No issues found — data looks clean![/green bold]\n")
        return

    # -- Column health overview -----------------------------------------------
    col_errors: Counter = Counter()
    col_warnings: Counter = Counter()
    for issue in result.issues:
        if issue.column:
            if issue.severity == Severity.ERROR:
                col_errors[issue.column] += 1
            elif issue.severity == Severity.WARNING:
                col_warnings[issue.column] += 1

    all_cols = sorted(set(list(col_errors.keys()) + list(col_warnings.keys())))
    if all_cols:
        col_table = Table(
            title="Column Health Overview",
            show_header=True, header_style="bold",
            show_lines=False, pad_edge=True, expand=False,
        )
        col_table.add_column("Column", width=18)
        col_table.add_column("Errors", justify="center", width=8)
        col_table.add_column("Warnings", justify="center", width=10)
        col_table.add_column("Health", width=24)

        for col in all_cols:
            errs = col_errors.get(col, 0)
            warns = col_warnings.get(col, 0)
            total_issues = errs + warns
            # Health bar: simple visual indicator
            if errs > 0:
                bar_color = "red"
                bar_fill = min(total_issues, 20)
                health = f"[{bar_color}]{'!' * bar_fill}[/{bar_color}]"
            elif warns > 0:
                bar_color = "yellow"
                bar_fill = min(total_issues, 20)
                health = f"[{bar_color}]{'~' * bar_fill}[/{bar_color}]"
            else:
                health = "[green]OK[/green]"

            col_table.add_row(
                col,
                f"[red]{errs}[/red]" if errs > 0 else "[dim]0[/dim]",
                f"[yellow]{warns}[/yellow]" if warns > 0 else "[dim]0[/dim]",
                health,
            )

        console.print(col_table)
        console.print()

    # -- Filter issues --------------------------------------------------------
    filtered = _filter_issues(result.issues, severity_filter, rule_filter)

    if not filtered:
        console.print(f"[dim]  No issues match the current filter (severity={severity_filter}, rule={rule_filter}).[/dim]\n")
        return

    # -- Issues table ---------------------------------------------------------
    display_issues = filtered[:limit] if limit else filtered
    truncated = len(filtered) - len(display_issues)

    table = Table(
        title=f"Issues ({len(display_issues)} of {len(filtered)} shown)",
        show_header=True, header_style="bold",
        show_lines=False, pad_edge=True,
    )
    table.add_column("Row", style="dim", width=6, justify="right")
    table.add_column("Column", width=16)
    table.add_column("Severity", width=9, justify="center")
    table.add_column("Rule", width=24)
    table.add_column("Message", min_width=40)

    severity_styles = {
        Severity.ERROR: "[red bold]ERROR[/red bold]",
        Severity.WARNING: "[yellow]WARN[/yellow]",
        Severity.INFO: "[cyan]INFO[/cyan]",
    }

    for issue in display_issues:
        table.add_row(
            str(issue.row) if issue.row else "—",
            issue.column or "—",
            severity_styles.get(issue.severity, str(issue.severity.value)),
            issue.rule,
            issue.message,
        )

    console.print(table)

    if truncated > 0:
        console.print(f"  [dim]... and {truncated:,} more issues. Use --limit to show more, or --format html for the full report.[/dim]")
    console.print()

    # -- Rule breakdown -------------------------------------------------------
    rule_counts: Counter = Counter()
    for issue in result.issues:
        rule_counts[issue.rule] += 1

    rule_table = Table(
        title="Issue Breakdown by Rule",
        show_header=True, header_style="bold",
        show_lines=False, pad_edge=True, expand=False,
    )
    rule_table.add_column("Rule", width=28)
    rule_table.add_column("Count", justify="right", width=8)
    rule_table.add_column("", width=30)

    max_count = max(rule_counts.values()) if rule_counts else 1
    for rule, count in sorted(rule_counts.items(), key=lambda x: -x[1]):
        bar_len = int((count / max_count) * 25)
        bar = "[blue]" + "█" * bar_len + "[/blue]"
        rule_table.add_row(rule, str(count), bar)

    console.print(rule_table)
    console.print()


def _filter_issues(
    issues: list[ValidationIssue],
    severity_filter: str = None,
    rule_filter: str = None,
) -> list[ValidationIssue]:
    """Filter issues by severity and/or rule name."""
    filtered = issues
    if severity_filter:
        sev = severity_filter.upper()
        filtered = [i for i in filtered if i.severity.value.upper() == sev]
    if rule_filter:
        rf = rule_filter.upper()
        filtered = [i for i in filtered if rf in i.rule.upper()]
    return filtered


# ---------------------------------------------------------------------------
# HTML report (interactive)
# ---------------------------------------------------------------------------

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Validation Report — {{ result.file_name }}</title>
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
  :root {
    --bg: #f5f6f8; --bg2: #ebedf2; --card: #ffffff; --text: #111827;
    --text2: #4b5563; --muted: #9ca3af; --border: #e5e7eb;
    --slate-900: #0f172a; --slate-800: #1e293b; --slate-700: #334155;
    --accent: #2563eb; --accent-light: #eff6ff;
    --red: #dc2626; --red-bg: #fef2f2; --red-border: #fecaca;
    --amber: #d97706; --amber-bg: #fffbeb; --amber-border: #fde68a;
    --green: #16a34a; --green-bg: #f0fdf4; --green-border: #bbf7d0;
    --blue: #2563eb; --blue-bg: #eff6ff; --blue-border: #bfdbfe;
  }
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body { font-family: 'IBM Plex Sans', system-ui, -apple-system, sans-serif;
         background: var(--bg); color: var(--text); line-height: 1.6;
         -webkit-font-smoothing: antialiased; }

  /* ── Header ──────────────────────────────────────────── */
  .header { background: var(--slate-900); padding: 2rem 2.5rem; color: #fff; }
  .header-inner { max-width: 1200px; margin: 0 auto; }
  .header-top { display: flex; justify-content: space-between; align-items: flex-start;
                flex-wrap: wrap; gap: 1rem; margin-bottom: 1.5rem; }
  .header h1 { font-size: 1.5rem; font-weight: 700; letter-spacing: -0.02em;
               margin-bottom: 0.25rem; }
  .header-sub { font-size: 0.8rem; color: rgba(255,255,255,0.5); font-weight: 400;
                letter-spacing: 0.02em; }
  .status-badge { font-size: 0.75rem; font-weight: 600; padding: 0.4rem 1.2rem;
                  border-radius: 4px; letter-spacing: 0.06em; text-transform: uppercase; }
  .status-pass { background: rgba(22,163,74,0.15); color: #4ade80; border: 1px solid rgba(22,163,74,0.3); }
  .status-fail { background: rgba(220,38,38,0.15); color: #fca5a5; border: 1px solid rgba(220,38,38,0.3); }

  .header-meta { display: flex; gap: 2.5rem; flex-wrap: wrap; font-size: 0.8rem;
                 color: rgba(255,255,255,0.45); padding-top: 1.25rem;
                 border-top: 1px solid rgba(255,255,255,0.08); }
  .header-meta span { display: flex; align-items: center; gap: 0.4rem; }
  .header-meta strong { color: rgba(255,255,255,0.7); font-weight: 500; }

  .container { max-width: 1200px; margin: 0 auto; padding: 2rem 2.5rem; }

  /* ── Metric cards ────────────────────────────────────── */
  .metrics { display: grid; grid-template-columns: repeat(5, 1fr); gap: 0.75rem;
             margin-bottom: 2rem; }
  @media (max-width: 900px) { .metrics { grid-template-columns: repeat(2, 1fr); } }
  .metric { background: var(--card); border: 1px solid var(--border); border-radius: 6px;
            padding: 1.25rem 1.5rem; }
  .metric .label { font-size: 0.68rem; font-weight: 600; text-transform: uppercase;
                   letter-spacing: 0.1em; color: var(--muted); margin-bottom: 0.5rem; }
  .metric .value { font-family: 'IBM Plex Mono', monospace; font-size: 2rem; font-weight: 700;
                   line-height: 1; }
  .metric.err .value { color: var(--red); }
  .metric.warn .value { color: var(--amber); }
  .metric.pass .value { color: var(--green); }

  .gauge { margin-top: 0.6rem; height: 6px; background: var(--bg2); border-radius: 3px;
           overflow: hidden; }
  .gauge-bar { height: 100%; border-radius: 3px; }

  /* ── Section headings ────────────────────────────────── */
  .section-label { font-size: 0.7rem; font-weight: 600; text-transform: uppercase;
                   letter-spacing: 0.1em; color: var(--slate-700); margin-bottom: 0.75rem;
                   padding-bottom: 0.5rem; border-bottom: 2px solid var(--slate-900);
                   display: inline-block; }
  .section-wrap { margin-bottom: 2rem; }

  /* ── Analytics panels ────────────────────────────────── */
  .panels { display: grid; grid-template-columns: 1fr 1fr; gap: 0.75rem; margin-bottom: 2rem; }
  @media (max-width: 768px) { .panels { grid-template-columns: 1fr; } }
  .panel { background: var(--card); border: 1px solid var(--border); border-radius: 6px;
           padding: 1.5rem; }
  .panel h3 { font-size: 0.7rem; font-weight: 600; text-transform: uppercase;
              letter-spacing: 0.08em; color: var(--muted); margin-bottom: 1rem; }

  /* Horizontal bars */
  .bar-row { display: flex; align-items: center; margin-bottom: 0.5rem; }
  .bar-label { width: 160px; font-size: 0.72rem; font-family: 'IBM Plex Mono', monospace;
               color: var(--text2); flex-shrink: 0; overflow: hidden;
               text-overflow: ellipsis; white-space: nowrap; }
  .bar-track { flex: 1; height: 16px; background: var(--bg2); border-radius: 3px;
               overflow: hidden; margin: 0 0.75rem; }
  .bar-fill { height: 100%; border-radius: 3px; }
  .bar-fill.sev-error { background: var(--red); }
  .bar-fill.sev-warning { background: var(--amber); }
  .bar-fill.sev-info { background: var(--blue); }
  .bar-count { width: 28px; font-size: 0.72rem; font-family: 'IBM Plex Mono', monospace;
               text-align: right; color: var(--muted); font-weight: 500; }

  /* Field health chips */
  .field-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr));
                gap: 0.5rem; }
  .field-chip { display: flex; align-items: center; gap: 0.6rem; padding: 0.6rem 0.85rem;
                border-radius: 4px; font-size: 0.8rem; font-weight: 500;
                border: 1px solid var(--border); background: var(--bg); }
  .field-dot { width: 8px; height: 8px; border-radius: 50%; flex-shrink: 0; }
  .field-dot.clean { background: var(--green); }
  .field-dot.warn { background: var(--amber); }
  .field-dot.bad { background: var(--red); }
  .field-chip .fname { flex: 1; font-family: 'IBM Plex Mono', monospace; font-size: 0.75rem; }
  .field-chip .fstat { font-family: 'IBM Plex Mono', monospace; font-size: 0.68rem;
                       color: var(--muted); }

  /* ── Toolbar ─────────────────────────────────────────── */
  .toolbar { display: flex; gap: 0.5rem; margin-bottom: 0.75rem; flex-wrap: wrap;
             align-items: center; }
  .search-input { background: var(--card); border: 1px solid var(--border);
                  color: var(--text); padding: 7px 12px 7px 34px; border-radius: 4px;
                  font-size: 0.82rem; width: 260px; outline: none;
                  font-family: 'IBM Plex Sans', sans-serif;
                  background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='14' height='14' viewBox='0 0 24 24' fill='none' stroke='%239ca3af' stroke-width='2'%3E%3Ccircle cx='11' cy='11' r='8'/%3E%3Cline x1='21' y1='21' x2='16.65' y2='16.65'/%3E%3C/svg%3E");
                  background-repeat: no-repeat; background-position: 10px center; }
  .search-input:focus { border-color: var(--accent); box-shadow: 0 0 0 2px rgba(37,99,235,0.1); }
  .filter-btn { background: var(--card); border: 1px solid var(--border);
                color: var(--text2); padding: 6px 14px; border-radius: 4px;
                font-size: 0.75rem; font-weight: 500; cursor: pointer;
                font-family: 'IBM Plex Sans', sans-serif; transition: all 0.12s; }
  .filter-btn:hover { border-color: var(--accent); color: var(--accent); }
  .filter-btn.active { background: var(--slate-900); color: #fff; border-color: var(--slate-900); }
  .record-count { color: var(--muted); font-size: 0.75rem; margin-left: auto;
                  font-family: 'IBM Plex Mono', monospace; }

  /* ── Data table ──────────────────────────────────────── */
  .table-container { background: var(--card); border: 1px solid var(--border);
                     border-radius: 6px; overflow: hidden; margin-bottom: 2rem; }
  table { width: 100%; border-collapse: collapse; }
  th { background: var(--slate-900); padding: 10px 14px; text-align: left;
       font-size: 0.68rem; font-weight: 600; text-transform: uppercase;
       letter-spacing: 0.08em; color: rgba(255,255,255,0.75);
       cursor: pointer; user-select: none; white-space: nowrap; }
  th:hover { color: #fff; }
  th .sort-ind { font-size: 0.6rem; margin-left: 4px; opacity: 0.5; }
  td { padding: 9px 14px; border-top: 1px solid var(--border); font-size: 0.82rem; }
  td:first-child { font-family: 'IBM Plex Mono', monospace; font-size: 0.78rem;
                   color: var(--muted); text-align: center; width: 55px; }
  td:nth-child(2) { font-weight: 500; }
  td:nth-child(4) { font-family: 'IBM Plex Mono', monospace; font-size: 0.74rem; }
  td:last-child { color: var(--text2); }
  tr:hover td { background: rgba(37,99,235,0.02); }
  tr.hidden { display: none; }
  tbody tr:nth-child(even) td { background: var(--bg); }
  tbody tr:nth-child(even):hover td { background: rgba(37,99,235,0.04); }

  /* Severity tags */
  .tag { display: inline-flex; align-items: center; gap: 4px; padding: 2px 8px;
         border-radius: 3px; font-size: 0.65rem; font-weight: 600;
         text-transform: uppercase; letter-spacing: 0.05em; }
  .tag::before { content: ''; width: 5px; height: 5px; border-radius: 50%; }
  .tag.error { background: var(--red-bg); color: var(--red); border: 1px solid var(--red-border); }
  .tag.error::before { background: var(--red); }
  .tag.warning { background: var(--amber-bg); color: var(--amber); border: 1px solid var(--amber-border); }
  .tag.warning::before { background: var(--amber); }
  .tag.info { background: var(--blue-bg); color: var(--blue); border: 1px solid var(--blue-border); }
  .tag.info::before { background: var(--blue); }

  /* Clean state */
  .clean-state { text-align: center; padding: 4rem 2rem; }
  .clean-state .icon { width: 48px; height: 48px; background: var(--green-bg);
                       border: 1px solid var(--green-border); border-radius: 50%;
                       display: inline-flex; align-items: center; justify-content: center;
                       margin-bottom: 1rem; }
  .clean-state .icon svg { color: var(--green); }
  .clean-state p { font-size: 1rem; color: var(--green); font-weight: 600; }
  .clean-state .sub { font-size: 0.82rem; color: var(--muted); font-weight: 400; margin-top: 0.3rem; }

  /* ── Footer ──────────────────────────────────────────── */
  footer { text-align: center; padding: 1.5rem 0; margin-top: 1rem;
           border-top: 1px solid var(--border); font-size: 0.72rem; color: var(--muted);
           letter-spacing: 0.02em; }

  /* Print */
  @media print {
    .header { print-color-adjust: exact; -webkit-print-color-adjust: exact; }
    .toolbar { display: none; }
    body { background: #fff; }
  }
</style>
</head>
<body>

<div class="header">
  <div class="header-inner">
    <div class="header-top">
      <div>
        <h1>Data Quality Validation Report</h1>
        <div class="header-sub">Automated validation engine &mdash; {{ result.validators_run }} rules applied</div>
      </div>
      <span class="status-badge {{ 'status-pass' if result.is_clean else 'status-fail' }}">
        {{ 'Validated' if result.is_clean else 'Issues Detected' }}
      </span>
    </div>
    <div class="header-meta">
      <span><strong>Source</strong>&ensp;{{ result.file_name }}</span>
      <span><strong>Generated</strong>&ensp;{{ timestamp }}</span>
      <span><strong>Records</strong>&ensp;{{ "{:,}".format(result.total_rows) }} rows &times; {{ result.total_columns }} fields</span>
    </div>
  </div>
</div>

<div class="container">

  <!-- Metrics -->
  <div class="metrics">
    <div class="metric">
      <div class="label">Total Records</div>
      <div class="value">{{ "{:,}".format(result.total_rows) }}</div>
    </div>
    <div class="metric">
      <div class="label">Fields Analyzed</div>
      <div class="value">{{ result.total_columns }}</div>
    </div>
    <div class="metric err">
      <div class="label">Errors</div>
      <div class="value">{{ result.error_count }}</div>
    </div>
    <div class="metric warn">
      <div class="label">Warnings</div>
      <div class="value">{{ result.warning_count }}</div>
    </div>
    <div class="metric pass">
      <div class="label">Pass Rate</div>
      <div class="value">{{ result.pass_rate }}%</div>
      <div class="gauge">
        <div class="gauge-bar" style="width:{{ result.pass_rate }}%;
             background: {% if result.pass_rate >= 90 %}var(--green){% elif result.pass_rate >= 60 %}var(--amber){% else %}var(--red){% endif %};"></div>
      </div>
    </div>
  </div>

  {% if result.issues %}

  <!-- Analytics -->
  <div class="section-wrap">
    <div class="section-label">Diagnostics</div>
    <div class="panels">
      <div class="panel">
        <h3>Issue Distribution by Rule</h3>
        {% for rule, count in rule_counts %}
        <div class="bar-row">
          <span class="bar-label" title="{{ rule }}">{{ rule }}</span>
          <div class="bar-track">
            <div class="bar-fill sev-{{ rule_severities[rule] }}"
                 style="width: {{ (count / max_rule_count * 100)|round|int }}%"></div>
          </div>
          <span class="bar-count">{{ count }}</span>
        </div>
        {% endfor %}
      </div>

      <div class="panel">
        <h3>Field Integrity</h3>
        <div class="field-grid">
          {% for col, status, errs, warns in column_health_detail %}
          <div class="field-chip">
            <span class="field-dot {{ status }}"></span>
            <span class="fname">{{ col }}</span>
            <span class="fstat">
              {% if status == 'clean' %}OK{% elif status == 'warn' %}{{ warns }}W{% else %}{{ errs }}E{{ ' / ' + warns|string + 'W' if warns > 0 else '' }}{% endif %}
            </span>
          </div>
          {% endfor %}
        </div>
      </div>
    </div>
  </div>

  <!-- Issue Log -->
  <div class="section-wrap">
    <div class="section-label">Issue Log</div>

    <div class="toolbar">
      <input type="text" class="search-input" id="searchBox"
             placeholder="Filter issues..." onkeyup="filterTable()">
      <button class="filter-btn active" onclick="toggleFilter(this, 'all')">All</button>
      <button class="filter-btn" onclick="toggleFilter(this, 'error')">Errors</button>
      <button class="filter-btn" onclick="toggleFilter(this, 'warning')">Warnings</button>
      <button class="filter-btn" onclick="toggleFilter(this, 'info')">Info</button>
      <span class="record-count" id="visibleCount">{{ result.issues|length }} records</span>
    </div>

    <div class="table-container">
    <table id="issuesTable">
      <thead>
        <tr>
          <th onclick="sortTable(0)" style="width:55px">Row <span class="sort-ind"></span></th>
          <th onclick="sortTable(1)" style="width:130px">Field <span class="sort-ind"></span></th>
          <th onclick="sortTable(2)" style="width:95px">Severity <span class="sort-ind"></span></th>
          <th onclick="sortTable(3)" style="width:180px">Rule <span class="sort-ind"></span></th>
          <th onclick="sortTable(4)">Description <span class="sort-ind"></span></th>
        </tr>
      </thead>
      <tbody>
      {% for issue in result.issues %}
        <tr data-severity="{{ issue.severity.value }}">
          <td>{{ issue.row if issue.row else '—' }}</td>
          <td>{{ issue.column if issue.column else '—' }}</td>
          <td><span class="tag {{ issue.severity.value }}">{{ issue.severity.value }}</span></td>
          <td>{{ issue.rule }}</td>
          <td>{{ issue.message }}</td>
        </tr>
      {% endfor %}
      </tbody>
    </table>
    </div>
  </div>

  {% else %}
  <div class="clean-state">
    <div class="icon">
      <svg width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"></polyline></svg>
    </div>
    <p>All records validated successfully</p>
    <div class="sub">No issues detected across {{ result.validators_run }} validation rules.</div>
  </div>
  {% endif %}

  <footer>
    Data Quality Validation Engine v1.0 &nbsp;&middot;&nbsp; Report generated {{ timestamp }}
  </footer>
</div>

<script>
let activeSeverity = 'all';

function toggleFilter(btn, severity) {
  document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  activeSeverity = severity;
  filterTable();
}

function filterTable() {
  const q = document.getElementById('searchBox').value.toLowerCase();
  const rows = document.querySelectorAll('#issuesTable tbody tr');
  let n = 0;
  rows.forEach(r => {
    const s = r.getAttribute('data-severity');
    const t = r.textContent.toLowerCase();
    const show = (activeSeverity === 'all' || s === activeSeverity) && (!q || t.includes(q));
    r.classList.toggle('hidden', !show);
    if (show) n++;
  });
  document.getElementById('visibleCount').textContent = n + ' record' + (n !== 1 ? 's' : '');
}

let sortCol = -1, sortAsc = true;
function sortTable(col) {
  const tbody = document.querySelector('#issuesTable tbody');
  const rows = Array.from(tbody.querySelectorAll('tr'));
  if (sortCol === col) sortAsc = !sortAsc; else { sortCol = col; sortAsc = true; }
  rows.sort((a, b) => {
    let va = a.cells[col].textContent.trim(), vb = b.cells[col].textContent.trim();
    if (col === 0) { va = parseInt(va)||0; vb = parseInt(vb)||0; return sortAsc ? va-vb : vb-va; }
    return sortAsc ? va.localeCompare(vb) : vb.localeCompare(va);
  });
  rows.forEach(r => tbody.appendChild(r));
  document.querySelectorAll('th .sort-ind').forEach((s,i) => {
    s.textContent = i === col ? (sortAsc ? ' ▲' : ' ▼') : '';
  });
}
</script>
</body>
</html>"""


def generate_html_report(result: ValidationResult, output_path: str | Path) -> Path:
    """Render the validation result as an interactive HTML file."""
    # Pre-compute chart data
    rule_counts_map: dict[str, int] = {}
    rule_severities: dict[str, str] = {}
    col_errors: Counter = Counter()
    col_warnings: Counter = Counter()

    for issue in result.issues:
        rule_counts_map[issue.rule] = rule_counts_map.get(issue.rule, 0) + 1
        if issue.severity == Severity.ERROR:
            rule_severities[issue.rule] = "error"
            if issue.column:
                col_errors[issue.column] += 1
        elif issue.severity == Severity.WARNING:
            rule_severities.setdefault(issue.rule, "warning")
            if issue.column:
                col_warnings[issue.column] += 1

    rule_counts = sorted(rule_counts_map.items(), key=lambda x: -x[1])
    max_rule_count = max(rule_counts_map.values()) if rule_counts_map else 1

    # Column health: all columns from the file
    all_columns = ["cusip", "issuer_name", "state", "issue_date", "maturity_date",
                   "par_value", "coupon_rate", "bond_type"]
    column_health_detail = []
    for col in all_columns:
        errs = col_errors.get(col, 0)
        warns = col_warnings.get(col, 0)
        if errs > 0:
            column_health_detail.append((col, "bad", errs, warns))
        elif warns > 0:
            column_health_detail.append((col, "warn", errs, warns))
        else:
            column_health_detail.append((col, "clean", 0, 0))

    template = Template(HTML_TEMPLATE)
    html = template.render(
        result=result,
        timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        rule_counts=rule_counts,
        max_rule_count=max_rule_count,
        rule_severities=rule_severities,
        column_health_detail=column_health_detail,
    )
    path = Path(output_path)
    path.write_text(html, encoding="utf-8")
    return path


# ---------------------------------------------------------------------------
# JSON report
# ---------------------------------------------------------------------------

def generate_json_report(result: ValidationResult, output_path: str | Path) -> Path:
    """Export validation results as structured JSON."""
    # Build column-level stats
    col_stats: dict[str, dict[str, int]] = {}
    for issue in result.issues:
        if issue.column:
            if issue.column not in col_stats:
                col_stats[issue.column] = {"errors": 0, "warnings": 0, "info": 0}
            col_stats[issue.column][issue.severity.value + "s" if issue.severity != Severity.INFO else "info"] += 1

    rule_counts: dict[str, int] = {}
    for issue in result.issues:
        rule_counts[issue.rule] = rule_counts.get(issue.rule, 0) + 1

    data = {
        "file_name": result.file_name,
        "generated_at": datetime.now().isoformat(),
        "summary": {
            "total_rows": result.total_rows,
            "total_columns": result.total_columns,
            "validators_run": result.validators_run,
            "errors": result.error_count,
            "warnings": result.warning_count,
            "info": result.info_count,
            "pass_rate": result.pass_rate,
        },
        "by_rule": dict(sorted(rule_counts.items(), key=lambda x: -x[1])),
        "by_column": col_stats,
        "issues": [
            {
                "row": i.row,
                "column": i.column,
                "severity": i.severity.value,
                "rule": i.rule,
                "message": i.message,
            }
            for i in result.issues
        ],
    }
    path = Path(output_path)
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")
    return path


# ---------------------------------------------------------------------------
# Excel (.xlsx) report
# ---------------------------------------------------------------------------

# Color constants
_NAVY = "0A2540"
_NAVY_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_NAVY_FILL = PatternFill("solid", fgColor=_NAVY)
_TEAL = "00838A"
_TEAL_FILL = PatternFill("solid", fgColor=_TEAL)
_LIGHT_GRAY = "F8F9FB"
_LIGHT_GRAY_FILL = PatternFill("solid", fgColor=_LIGHT_GRAY)
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
_RED_FILL = PatternFill("solid", fgColor="FDF0F2")
_RED_FONT = Font(name="Arial", bold=True, color="D03050", size=10)
_YELLOW_FILL = PatternFill("solid", fgColor="FEF9EE")
_YELLOW_FONT = Font(name="Arial", bold=True, color="C27800", size=10)
_GREEN_FILL = PatternFill("solid", fgColor="EDF9F0")
_GREEN_FONT = Font(name="Arial", bold=True, color="0E7C3A", size=10)
_BLUE_FILL = PatternFill("solid", fgColor="EEF4FC")
_BLUE_FONT = Font(name="Arial", bold=True, color="1A6DD4", size=10)
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="DFE4EB"),
)
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_TITLE_FONT = Font(name="Arial", bold=True, color=_NAVY, size=14)
_SUBTITLE_FONT = Font(name="Arial", color="8896A6", size=9)
_BODY_FONT = Font(name="Arial", color="1A2332", size=10)
_MONO_FONT = Font(name="Consolas", color="4A5568", size=9)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _NAVY_FILL
        cell.alignment = _CENTER


def _auto_width(ws, min_w=10, max_w=50):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value:
                lengths.append(len(str(cell.value)))
        if lengths:
            best = min(max(max(lengths) + 2, min_w), max_w)
            ws.column_dimensions[col_letter].width = best


def generate_xlsx_report(result: ValidationResult, output_path: str | Path) -> Path:
    """Generate a professionally formatted Excel validation report."""
    wb = Workbook()

    # ── Sheet 1: Summary Dashboard ──────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = _NAVY

    # Set fixed column widths (A-D used for data, no extra columns)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18

    # ── Title block (rows 1-3) ──
    ws.merge_cells("A1:D1")
    ws["A1"] = "Data Quality Report"
    ws["A1"].font = Font(name="Arial", bold=True, color=_NAVY, size=16)

    ws.merge_cells("A2:D2")
    ws["A2"] = "Automated Data Validation Engine"
    ws["A2"].font = _SUBTITLE_FONT

    ws.merge_cells("A3:D3")
    ws["A3"] = f"Source: {result.file_name}  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  Validators: {result.validators_run}"
    ws["A3"].font = _SUBTITLE_FONT

    # ── Metrics row (rows 5-6) ──
    metric_data = [
        ("Records Scanned", result.total_rows),
        ("Errors", result.error_count),
        ("Warnings", result.warning_count),
        ("Pass Rate", f"{result.pass_rate}%"),
    ]

    for col_idx, (label, value) in enumerate(metric_data, 1):
        label_cell = ws.cell(row=5, column=col_idx, value=label)
        label_cell.font = Font(name="Arial", bold=True, color=_NAVY, size=9)
        label_cell.alignment = _CENTER
        label_cell.fill = PatternFill("solid", fgColor="E8ECF1")

        val_cell = ws.cell(row=6, column=col_idx, value=value)
        val_cell.font = Font(name="Arial", bold=True, size=18)
        val_cell.alignment = _CENTER

        if label == "Errors":
            val_cell.font = Font(name="Arial", bold=True, color="D03050", size=18)
        elif label == "Warnings":
            val_cell.font = Font(name="Arial", bold=True, color="C27800", size=18)
        elif label == "Pass Rate":
            if result.pass_rate >= 90:
                val_cell.font = Font(name="Arial", bold=True, color="0E7C3A", size=18)
            elif result.pass_rate >= 60:
                val_cell.font = Font(name="Arial", bold=True, color="C27800", size=18)
            else:
                val_cell.font = Font(name="Arial", bold=True, color="D03050", size=18)

    # Status indicator
    ws.merge_cells("A7:D7")
    status_cell = ws.cell(row=7, column=1)
    if result.is_clean:
        status_cell.value = "STATUS: VALIDATED"
        status_cell.font = Font(name="Arial", bold=True, color="0E7C3A", size=11)
        status_cell.fill = _GREEN_FILL
    else:
        status_cell.value = "STATUS: ISSUES FOUND"
        status_cell.font = Font(name="Arial", bold=True, color="D03050", size=11)
        status_cell.fill = _RED_FILL
    status_cell.alignment = _CENTER

    # ── Rule breakdown table (row 9+) ──
    row = 9
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="ISSUE DISTRIBUTION BY RULE").font = Font(name="Arial", bold=True, color=_NAVY, size=11)

    row = 10
    for c, header in enumerate(["Rule", "Count", "Severity", ""], 1):
        cell = ws.cell(row=row, column=c, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _NAVY_FILL
        cell.alignment = _CENTER

    rule_counts: Counter = Counter()
    rule_sev: dict = {}
    for issue in result.issues:
        rule_counts[issue.rule] += 1
        if issue.severity == Severity.ERROR:
            rule_sev[issue.rule] = "ERROR"
        elif issue.rule not in rule_sev:
            rule_sev[issue.rule] = issue.severity.value.upper()

    max_count = max(rule_counts.values()) if rule_counts else 1
    for rule, count in sorted(rule_counts.items(), key=lambda x: -x[1]):
        row += 1
        ws.cell(row=row, column=1, value=rule).font = _MONO_FONT
        ws.cell(row=row, column=2, value=count).alignment = _CENTER
        sev = rule_sev.get(rule, "INFO")
        sev_cell = ws.cell(row=row, column=3, value=sev)
        sev_cell.alignment = _CENTER
        if sev == "ERROR":
            sev_cell.font = _RED_FONT
            sev_cell.fill = _RED_FILL
        elif sev == "WARNING":
            sev_cell.font = _YELLOW_FONT
            sev_cell.fill = _YELLOW_FILL
        else:
            sev_cell.font = _BLUE_FONT
            sev_cell.fill = _BLUE_FILL
        # Visual bar in column D
        bar_len = int((count / max_count) * 15)
        bar_cell = ws.cell(row=row, column=4, value="\u2588" * bar_len)
        bar_cell.font = Font(name="Arial", color=_TEAL, size=10)

    rule_table_end = row

    # ── Field integrity table (below rules) ──
    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="FIELD INTEGRITY STATUS").font = Font(name="Arial", bold=True, color=_NAVY, size=11)

    row += 1
    for c, header in enumerate(["Field", "Errors", "Warnings", "Status"], 1):
        cell = ws.cell(row=row, column=c, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _NAVY_FILL
        cell.alignment = _CENTER

    col_errors: Counter = Counter()
    col_warnings: Counter = Counter()
    for issue in result.issues:
        if issue.column:
            if issue.severity == Severity.ERROR:
                col_errors[issue.column] += 1
            elif issue.severity == Severity.WARNING:
                col_warnings[issue.column] += 1

    all_columns = ["cusip", "issuer_name", "state", "issue_date", "maturity_date",
                   "par_value", "coupon_rate", "bond_type"]
    for col_name in all_columns:
        row += 1
        ws.cell(row=row, column=1, value=col_name).font = _MONO_FONT

        errs = col_errors.get(col_name, 0)
        warns = col_warnings.get(col_name, 0)

        err_cell = ws.cell(row=row, column=2, value=errs)
        err_cell.alignment = _CENTER
        err_cell.font = _RED_FONT if errs > 0 else _BODY_FONT

        warn_cell = ws.cell(row=row, column=3, value=warns)
        warn_cell.alignment = _CENTER
        warn_cell.font = _YELLOW_FONT if warns > 0 else _BODY_FONT

        status = "OK" if errs == 0 and warns == 0 else ("ERRORS" if errs > 0 else "WARNINGS")
        status_cell = ws.cell(row=row, column=4, value=status)
        status_cell.alignment = _CENTER
        if status == "OK":
            status_cell.font = _GREEN_FONT
            status_cell.fill = _GREEN_FILL
        elif status == "ERRORS":
            status_cell.font = _RED_FONT
            status_cell.fill = _RED_FILL
        else:
            status_cell.font = _YELLOW_FONT
            status_cell.fill = _YELLOW_FILL

        # Alternate row shading
        if row % 2 == 0:
            for c in range(1, 5):
                if ws.cell(row=row, column=c).fill == PatternFill():
                    ws.cell(row=row, column=c).fill = _LIGHT_GRAY_FILL

    # ── Bar chart (below everything, full width) ──
    if rule_counts:
        row += 2
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Issues by Rule"
        chart.y_axis.title = None
        chart.x_axis.title = None
        chart.legend = None
        chart.width = 28
        chart.height = max(10, len(rule_counts) * 1.5)
        data_ref = Reference(ws, min_col=2, min_row=10, max_row=rule_table_end)
        cats_ref = Reference(ws, min_col=1, min_row=11, max_row=rule_table_end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill = _TEAL
        ws.add_chart(chart, f"A{row}")

    # ── Sheet 2: Issue Log ──────────────────────────────────────────────
    ws_issues = wb.create_sheet("Issue Log")
    ws_issues.sheet_properties.tabColor = "D03050" if result.error_count > 0 else "0E7C3A"

    headers = ["Row", "Field", "Severity", "Rule", "Description"]
    for c, h in enumerate(headers, 1):
        cell = ws_issues.cell(row=1, column=c, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _NAVY_FILL
        cell.alignment = _CENTER

    ws_issues.auto_filter.ref = f"A1:E{max(2, len(result.issues) + 1)}"

    for r, issue in enumerate(result.issues, 2):
        ws_issues.cell(row=r, column=1, value=issue.row or "—").font = _MONO_FONT
        ws_issues.cell(row=r, column=1).alignment = _CENTER
        ws_issues.cell(row=r, column=2, value=issue.column or "—").font = Font(name="Arial", bold=True, size=10)
        sev_cell = ws_issues.cell(row=r, column=3, value=issue.severity.value.upper())
        sev_cell.alignment = _CENTER
        if issue.severity == Severity.ERROR:
            sev_cell.font = _RED_FONT
            sev_cell.fill = _RED_FILL
        elif issue.severity == Severity.WARNING:
            sev_cell.font = _YELLOW_FONT
            sev_cell.fill = _YELLOW_FILL
        else:
            sev_cell.font = _BLUE_FONT
            sev_cell.fill = _BLUE_FILL
        ws_issues.cell(row=r, column=4, value=issue.rule).font = _MONO_FONT
        ws_issues.cell(row=r, column=5, value=issue.message).font = _BODY_FONT
        ws_issues.cell(row=r, column=5).alignment = _WRAP

        if r % 2 == 0:
            for c in range(1, 6):
                if ws_issues.cell(row=r, column=c).fill == PatternFill():
                    ws_issues.cell(row=r, column=c).fill = _LIGHT_GRAY_FILL

    ws_issues.column_dimensions["A"].width = 8
    ws_issues.column_dimensions["B"].width = 16
    ws_issues.column_dimensions["C"].width = 12
    ws_issues.column_dimensions["D"].width = 26
    ws_issues.column_dimensions["E"].width = 60
    ws_issues.freeze_panes = "A2"

    # ── Save ────────────────────────────────────────────────────────────
    path = Path(output_path)
    wb.save(path)
    return path
